"""
elevata - Metadata-driven Data Platform Framework
Copyright © 2025-2026 Ilona Tag

This file is part of elevata.

elevata is free software: you can redistribute it and/or modify
it under the terms of the GNU Affero General Public License as
published by the Free Software Foundation, either version 3 of
the License, or (at your option) any later version.

elevata is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
GNU Affero General Public License for more details.

You should have received a copy of the GNU Affero General Public License
along with elevata. If not, see <https://www.gnu.org/licenses/>.

Contact: <https://github.com/elevata-labs/elevata>.
"""

from __future__ import annotations

import time
import datetime
import os
import csv
import json
from io import BytesIO
import urllib.parse
import urllib.request
from pathlib import Path

from sqlalchemy import text

from metadata.rendering.dialects.dialect_factory import get_active_dialect
from metadata.ingestion.connectors import engine_for_source_system
from metadata.rendering.builder import qualify_source_filter
from metadata.rendering.placeholders import (
  resolve_delta_cutoff_for_source_dataset,
  apply_delta_cutoff_placeholder,
)
from metadata.materialization.logging import (
  ensure_load_run_log_table,
  build_load_run_log_row,
)
from metadata.ingestion.landing import render_param_insert_sql, land_raw_json_records
from metadata.ingestion.normalization import (
  normalize_column_name,
  normalize_records_keep_payload,
)


META_SCHEMA = "meta"


def _now_utc():
  return datetime.datetime.utcnow()


def _read_bytes(uri: str, *, max_bytes: int = 10_000_000) -> bytes:
  """
  Best-effort reader for file-like URIs.
  Supports:
    - http(s)://...
    - file://...
    - local paths
  """
  uri = os.path.expandvars(uri or "")
  p = urllib.parse.urlparse(uri or "")
  if p.scheme in ("http", "https"):
    with urllib.request.urlopen(uri, timeout=60) as resp:
      data = resp.read(max_bytes + 1)
    return data[:max_bytes]

  if uri.startswith("file://"):
    uri = uri[len("file://"):]
    # Windows file URI fix: file:///C:/... becomes /C:/... after stripping.
    # Convert /C:/... -> C:/... so os.path.exists works on Windows.
    if uri and len(uri) >= 4 and uri[0] == "/" and uri[2] == ":" and uri[3] == "/":
      # Example: /C:/temp/x.csv -> C:/temp/x.csv
      uri = uri[1:]

  if not os.path.exists(uri):
    raise ValueError(f"File not found: {uri}")

  with open(uri, "rb") as f:
    return f.read(max_bytes)


def _suffix_from_uri(uri: str) -> str:
  """
  Determine file suffix from URI path.
  Works for http(s):// and file:// and local paths.
  """
  uri = os.path.expandvars(uri or "")
  p = urllib.parse.urlparse(uri or "")
  path = p.path if p.scheme in ("http", "https") else uri
  if isinstance(path, str) and path.startswith("file://"):
    path = path[len("file://"):]
  return Path(path).suffix.lower()


def _load_file_records(
  uri: str,
  *,
  file_type: str | None = None,
  delimiter: str | None = None,
  quotechar: str | None = None,
  encoding: str | None = None,
) -> list[dict]:
  """
  Load file content into list of dict records.
  Supports:
    - .json  (JSON array)
    - .jsonl (NDJSON)
    - .csv
    - .xlsx/.xlsm (Excel)
  """
  ft = (file_type or "").strip().lower()
  suffix = _suffix_from_uri(uri)
  raw = _read_bytes(uri)

  if ft == "excel" or suffix in (".xlsx", ".xlsm"):
    # Default Excel behavior: first sheet, header row 1
    return _load_excel_records_from_bytes(raw)

  if ft == "json" or suffix == ".json":
    data = json.loads(raw.decode("utf-8", errors="replace"))
    if isinstance(data, list):
      return normalize_records_keep_payload([r for r in data if isinstance(r, dict)])    
    raise ValueError("JSON file must contain an array of objects.")

  if ft in ("jsonl", "ndjson") or suffix in (".jsonl", ".ndjson"):
    rows = []
    txt = raw.decode((encoding or "utf-8"), errors="replace")

    for line in txt.splitlines():
      line = (line or "").strip()
      if not line:
        continue
      obj = json.loads(line)
      if isinstance(obj, dict):
        rows.append(obj)
    return normalize_records_keep_payload(rows)  

  if ft == "csv" or suffix == ".csv":    
    rows = []
    txt = raw.decode((encoding or "utf-8"), errors="replace")
    reader = csv.DictReader(
      txt.splitlines(),
      delimiter=(delimiter or ","),
      quotechar=(quotechar or '"'),
    )

    for r in reader:
      rows.append(dict(r))
    return normalize_records_keep_payload(rows)

  raise ValueError(f"Unsupported file type: {suffix or ft or '<?>'}")


def _local_path_from_uri(uri: str) -> str:
  """
  Convert file:// URIs to local paths. Keeps plain local paths unchanged.
  Supports:
    - file:///C:/path/file.parquet  -> C:/path/file.parquet
    - file://C:/path/file.parquet   -> C:/path/file.parquet
    - C:/path/file.parquet          -> C:/path/file.parquet
  """
  u = (uri or "").strip()

  if u.startswith("file://"):
    u = u[len("file://"):]  # may leave "/C:/..." on Windows

  # Windows drive-letter normalization: "/C:/..." -> "C:/..."
  if len(u) >= 3 and u[0] == "/" and u[2] == ":" and u[1].isalpha():
    u = u[1:]

  return u


def _iter_parquet_record_chunks(path: str, *, chunk_size: int) -> "list[list[dict]]":
  """
  Yield chunks of rows from a Parquet file as list[dict].
  Uses pyarrow ParquetFile.iter_batches for efficient batch reads.
  """
  import pyarrow.parquet as pq

  pf = pq.ParquetFile(path)
  for batch in pf.iter_batches(batch_size=chunk_size):
    data = batch.to_pydict()  # dict[col] -> list(values)
    cols = list(data.keys())
    if not cols:
      continue
    n = len(data[cols[0]])
    out = []
    for i in range(n):
      out.append({c: data[c][i] for c in cols})
    if out:
      yield out


def _excel_cell_to_jsonable(value):
  """
  Convert Excel cell values into JSON-friendly values.
  """
  if isinstance(value, (datetime.datetime, datetime.date)):
    return value.isoformat()
  return value


def _load_excel_records_from_bytes(
  raw: bytes,
  *,
  sheet_name: str | None = None,
  sheet_index: int | None = None,
  header_row: int = 1,
  max_rows: int | None = None,
) -> list[dict]:
  """
  Load an Excel file into list[dict].
  - header_row is 1-based.
  - If sheet_name and sheet_index are both None, uses the first sheet.
  """
  from openpyxl import load_workbook

  wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)

  if sheet_name is not None:
    if sheet_name not in wb.sheetnames:
      raise ValueError(f"Excel sheet not found: {sheet_name!r}. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
  else:
    idx = sheet_index if sheet_index is not None else 0
    if idx < 0 or idx >= len(wb.sheetnames):
      raise ValueError(f"Excel sheet_index out of range: {idx}. Available: {wb.sheetnames}")
    ws = wb[wb.sheetnames[idx]]

  if header_row < 1:
    raise ValueError("header_row must be >= 1")

  # Read header
  header = None
  for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if i == header_row:
      header = [str(c).strip() if c is not None and str(c).strip() else f"col_{j+1}" for j, c in enumerate(row)]
      break

  if header is None:
    return []

  # Read data rows
  records = []
  data_start = header_row + 1
  data_count = 0

  for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if i < data_start:
      continue

    values = list(row)
    # Stop condition is intentionally conservative: we do not auto-stop on empty rows
    # because some sheets have blanks. max_rows is the safe limiter.

    rec = {}
    for j, col in enumerate(header):
      v = values[j] if j < len(values) else None
      rec[col] = _excel_cell_to_jsonable(v)
    records.append(rec)

    data_count += 1
    if max_rows is not None and data_count >= max_rows:
      break

  return records


def ingest_raw_relational(
  *,
  source_dataset,
  td,
  target_system,
  dialect,
  profile,
  batch_run_id: str,
  load_run_id: str,
  chunk_size: int = 5000,
  **kwargs,
) -> dict[str, object]:
  """
  Relational/native RAW ingestion (Source DB -> RAW landing table).
  With externally provided td/load_run_id/batch_run_id to keep orchestration stable.
  """
  target_dialect = dialect
  target_engine = target_dialect.get_execution_engine(target_system)

  source_sa_engine = engine_for_source_system(
    system_type=source_dataset.source_system.type,
    short_name=source_dataset.source_system.short_name,
  )
  source_dialect = get_active_dialect(source_sa_engine.dialect.name)

  # Determine columns to extract (integrated source columns, ordered)
  src_cols_qs = source_dataset.source_columns.filter(integrate=True).order_by("ordinal_position")
  src_col_names = [c.source_column_name for c in src_cols_qs]

  if not src_col_names:
    return {
      "status": "skipped",
      "reason": "no_integrated_columns",
      "load_run_id": load_run_id,
      "batch_run_id": batch_run_id,
    }

  # Build SELECT against source
  src_schema = source_dataset.schema_name
  src_table = source_dataset.source_dataset_name

  src_select_cols = ", ".join(f"s.{source_dialect.render_identifier(c)}" for c in src_col_names)
  src_from = source_dialect.render_table_identifier(src_schema, src_table)

  # Use a stable alias for filtering
  src_sql = f"SELECT {src_select_cols} FROM {src_from} AS s"

  static_filter = (getattr(source_dataset, "static_filter", None) or "").strip()
  increment_filter = (getattr(source_dataset, "increment_filter", None) or "").strip()

  where_parts = []

  # Static filter applies to ingestion only (RAW extraction)
  if static_filter:
    where_parts.append(f"({qualify_source_filter(source_dataset, static_filter, source_alias='s')})")

  # Increment filter applies to ingestion only when dataset is incremental
  apply_increment = bool(getattr(source_dataset, "incremental", False) and increment_filter)
  if apply_increment:
    where_parts.append(f"({qualify_source_filter(source_dataset, increment_filter, source_alias='s')})")

  if where_parts:
    src_sql += " WHERE " + " AND ".join(where_parts)

  src_sql += ";"

  # Replace {{DELTA_CUTOFF}} for incremental extraction on the SOURCE side
  cutoff = None
  if apply_increment and "{{DELTA_CUTOFF" in src_sql:
    cutoff = resolve_delta_cutoff_for_source_dataset(
      source_dataset=source_dataset,
      profile=profile,
      now_ts=_now_utc(),
    )
    if cutoff is None:
      raise ValueError(
        f"increment_filter uses {{DELTA_CUTOFF}} but no active increment policy exists "
        f"for SourceDataset={source_dataset} in environment '{getattr(profile, 'name', None)}'."
      )

    # IMPORTANT: render literal using SOURCE dialect
    src_sql = apply_delta_cutoff_placeholder(
      src_sql,
      dialect=source_dialect,
      delta_cutoff=cutoff,
    )

  # Build INSERT into RAW (DuckDB uses ? placeholders)
  # Target columns are derived from generated TargetColumns in RAW dataset.
  tgt_cols_qs = td.target_columns.filter(active=True).order_by("ordinal_position")
  tgt_cols = list(tgt_cols_qs)

  TECH_ROLES = {"payload", "load_run_id", "loaded_at"}
  TECH_NAMES = {"payload", "load_run_id", "loaded_at"}

  def _is_tech(col) -> bool:
    role = (col.system_role or "").strip()
    return (role in TECH_ROLES) or (col.target_column_name in TECH_NAMES)

  business_cols = [c for c in tgt_cols if not _is_tech(c)]
  tech_cols = [c for c in tgt_cols if _is_tech(c)]

  business_col_names = [c.target_column_name for c in business_cols]
  tech_col_names_found = [c.target_column_name for c in tech_cols]

  if len(src_col_names) != len(business_col_names):
    raise ValueError(
      f"Column mismatch for RAW ingestion: source has {len(src_col_names)} integrated columns "
      f"but target has {len(business_col_names)} business columns "
      f"(plus {len(tech_col_names_found)} technical columns)."
    )

  tech_col_names = [n for n in ("payload", "load_run_id", "loaded_at") if n in tech_col_names_found]

  insert_cols = business_col_names + tech_col_names

  insert_sql = render_param_insert_sql(
    dialect=target_dialect,
    schema_name=td.target_schema.schema_name,
    table_name=td.target_dataset_name,
    target_columns=insert_cols,
  )

  started_at = _now_utc()
  loaded_at = started_at
  t0 = time.time()

  rows_affected = 0

  try:
    # Ensure meta logging table exists
    ensure_load_run_log_table(
      engine=target_engine,
      dialect=target_dialect,
      meta_schema=META_SCHEMA,
      auto_provision=True,
    )

    # Ensure RAW schema/table exist
    target_engine.execute(target_dialect.render_create_schema_if_not_exists(td.target_schema.schema_name))
    # RAW is a landing area and expected to evolve with the source schema.
    # For full ingests we prefer DROP+CREATE to avoid stale schemas (missing new columns).
    if hasattr(target_dialect, "render_drop_table_if_exists"):
      is_raw = (getattr(getattr(td, "target_schema", None), "short_name", None) or "").lower() == "raw"
      drop_sql = target_dialect.render_drop_table_if_exists(
        schema=td.target_schema.schema_name,
        table=td.target_dataset_name,
        cascade=is_raw,
      )
      if drop_sql:
        target_engine.execute(drop_sql)
    target_engine.execute(target_dialect.render_create_table_if_not_exists(td))

    # Truncate RAW table (RAW is always materialized as table)
    target_engine.execute(
      target_dialect.render_truncate_table(
        schema=td.target_schema.schema_name,
        table=td.target_dataset_name,
      )
    )

    # Stream source rows and insert into RAW in chunks
    with source_sa_engine.connect() as conn:
      result = conn.execute(text(src_sql))
      chunk = []

      for row in result:
        chunk.append(row)
        if len(chunk) >= chunk_size:
          params = []
          for r in chunk:
            values = list(tuple(r))
            for tech_name in tech_col_names:
              if tech_name == "load_run_id":
                values.append(load_run_id)
              elif tech_name == "loaded_at":
                values.append(loaded_at)
              else:
                values.append(None)
            params.append(tuple(values))

          target_engine.execute_many(insert_sql, params)
          rows_affected += len(chunk)
          chunk = []

      if chunk:
        params = []
        for r in chunk:
          values = list(tuple(r))
          for tech_name in tech_col_names:
            if tech_name == "load_run_id":
              values.append(load_run_id)
            elif tech_name == "loaded_at":
              values.append(loaded_at)
            else:
              values.append(None)
          params.append(tuple(values))

        target_engine.execute_many(insert_sql, params)
        rows_affected += len(chunk)

    finished_at = _now_utc()
    exec_ms = (time.time() - t0) * 1000.0

    summary = {
      "schema": td.target_schema.short_name,
      "dataset": td.target_dataset_name,
      "target_dataset_id": td.id,
      "mode": ("incremental" if apply_increment else "full"),
      "handle_deletes": False,
      "historize": False,
    }

    values = build_load_run_log_row(
      batch_run_id=batch_run_id,
      load_run_id=load_run_id,
      target_schema=td.target_schema.short_name,
      target_dataset=td.target_dataset_name,
      target_system=target_system.short_name,
      profile=profile.name,
      run_kind="ingestion",
      source_system=str(getattr(getattr(source_dataset, "source_system", None), "short_name", None) or ""),
      source_dataset=str(getattr(source_dataset, "source_dataset_name", None) or ""),
      source_object=f"{src_schema}.{src_table}",
      ingest_mode=("incremental" if apply_increment else "full"),
      delta_cutoff=cutoff,
      rows_extracted=rows_affected,
      chunk_size=int(chunk_size),
      mode=str(summary.get("mode") or "full"),
      handle_deletes=bool(summary.get("handle_deletes") or False),
      historize=bool(summary.get("historize") or False),
      started_at=started_at,
      finished_at=finished_at,
      render_ms=0,
      execution_ms=exec_ms,
      sql_length=0,
      rows_affected=rows_affected,
      status="success",
      error_message=None,
      attempt_no=1,
      status_reason=None,
      blocked_by=None,
    )
    log_sql = target_dialect.render_insert_load_run_log(meta_schema=META_SCHEMA, values=values)
    if log_sql:
      target_engine.execute(log_sql)

    return {
      "status": "success",
      "rows_affected": rows_affected,
      "load_run_id": load_run_id,
      "batch_run_id": batch_run_id,
      "target_dataset": td.target_dataset_name,
      "source_sql": src_sql,
    }

  except Exception as e:
    err = str(e)
    finished_at = _now_utc()
    exec_ms = (time.time() - t0) * 1000.0

    summary = {
      "schema": td.target_schema.short_name,
      "dataset": td.target_dataset_name,
      "target_dataset_id": td.id,
      "mode": ("incremental" if apply_increment else "full"),
      "handle_deletes": False,
      "historize": False,
    }

    try:
      values = build_load_run_log_row(
        batch_run_id=batch_run_id,
        load_run_id=load_run_id,
        target_schema=td.target_schema.short_name,
        target_dataset=td.target_dataset_name,
        target_system=target_system.short_name,
        profile=profile.name,
        run_kind="ingestion",
        source_system=str(getattr(getattr(source_dataset, "source_system", None), "short_name", None) or ""),
        source_dataset=str(getattr(source_dataset, "source_dataset_name", None) or ""),
        source_object=f"{src_schema}.{src_table}",
        ingest_mode=("incremental" if apply_increment else "full"),
        delta_cutoff=cutoff,
        rows_extracted=rows_affected,
        chunk_size=int(chunk_size),
        mode=str(summary.get("mode") or "full"),
        handle_deletes=bool(summary.get("handle_deletes") or False),
        historize=bool(summary.get("historize") or False),
        started_at=started_at,
        finished_at=finished_at,
        render_ms=0,
        execution_ms=exec_ms,
        sql_length=0,
        rows_affected=rows_affected,
        status="error",
        error_message=(err or "")[:1000],
        attempt_no=1,
        status_reason=None,
        blocked_by=None,
      )
      log_sql = target_dialect.render_insert_load_run_log(meta_schema=META_SCHEMA, values=values)
      if log_sql:
        target_engine.execute(log_sql)
    except Exception:
      # Logging must never mask the original failure
      pass

    raise


def ingest_raw_file(
  *,
  source_dataset,
  td,
  target_system,
  dialect,
  profile,
  batch_run_id: str,
  load_run_id: str,
  meta_schema: str = "meta",
  chunk_size: int = 10_000,
  file_type: str | None = None,
):
  """
  File ingestion (JSON array / JSONL / CSV / Parquet).
  """
  uri = source_dataset.ingestion_config.get("uri")
  if not uri:
    raise ValueError("File ingestion requires ingestion_config.uri")

  ft = (file_type or "").strip().lower()
  uri_s = os.path.expandvars(str(uri))

  # --- Guard: file_type must match System.type if explicitly set ---
  system_type = str(source_dataset.source_system.type).lower()
  if ft:
    if system_type in ("csv", "json", "jsonl", "parquet", "excel"):
      if ft != system_type:
        raise ValueError(
          f"Inconsistent file configuration: System.type='{system_type}' "
          f"but ingestion_config.file_type='{ft}'. They must match."
        )

  # Use dialect-owned execution engine (consistent with relational ingestion).
  # This avoids relying on SQLAlchemy Engine semantics for landing.
  target_engine = dialect.get_execution_engine(target_system)

  rows_extracted = 0
  landing = None
  rows_inserted_total = 0

  if ft == "parquet":
    path = _local_path_from_uri(uri_s)
    if not os.path.exists(path):
      raise ValueError(f"File not found: {path}")

    started_at = datetime.datetime.now(datetime.timezone.utc)

    # Ensure log table exists once
    ensure_load_run_log_table(
      engine=target_engine,
      dialect=dialect,
      meta_schema=meta_schema,
      auto_provision=True,
    )

    first = True
    for chunk in _iter_parquet_record_chunks(path, chunk_size=chunk_size):

      rows_extracted += len(chunk)
      landing_part = land_raw_json_records(
        target_engine=target_engine,
        target_dialect=dialect,
        td=td,
        records=chunk,
        batch_run_id=batch_run_id,
        load_run_id=load_run_id,
        target_system=target_system,
        profile=profile,
        meta_schema=meta_schema,
        source_system_short_name=str(source_dataset.source_system.short_name),
        source_dataset_name=str(source_dataset.source_dataset_name),
        source_object=uri_s,
        ingest_mode=(file_type or "file"),
        chunk_size=chunk_size,
        source_dataset=source_dataset,
        strict=False,
        rebuild=first,
        write_run_log=False,
      )
      rows_inserted_total += int((landing_part or {}).get("rows_inserted") or 0)
      landing = landing_part
      first = False

    # Ensure returned landing reflects total inserted rows across chunks
    if rows_inserted_total > 0:
      landing = {**(landing or {}), "rows_inserted": rows_inserted_total}

    if rows_extracted == 0:
      return {
        "rows_extracted": rows_extracted,
        "landing": {**(landing or {}), "rows_inserted": rows_inserted_total},
      }

    finished_at = datetime.datetime.now(datetime.timezone.utc)

    # Write exactly one run log row (best-effort)
    try:
      values = build_load_run_log_row(
        batch_run_id=batch_run_id,
        load_run_id=load_run_id,
        target_schema=td.target_schema.short_name,
        target_dataset=td.target_dataset_name,
        target_system=target_system.short_name,
        profile=profile.name,
        run_kind="ingestion",
        source_system=str(source_dataset.source_system.short_name),
        source_dataset=str(source_dataset.source_dataset_name),
        source_object=uri_s,
        ingest_mode=(file_type or "file"),
        delta_cutoff=None,
        rows_extracted=rows_extracted,
        chunk_size=int(chunk_size),
        mode="full",
        handle_deletes=False,
        historize=False,
        started_at=started_at,
        finished_at=finished_at,
        render_ms=0.0,
        execution_ms=0.0,
        sql_length=0,
        rows_affected=rows_extracted,
        status="success",
        error_message=None,
        attempt_no=1,
      )
      sql = dialect.render_insert_load_run_log(meta_schema=meta_schema, values=values)
      if sql:
        target_engine.execute(sql)
    except Exception:
      pass
  else:
    # For Excel we allow extra ingestion_config options
    if system_type == "excel" or ft == "excel" or Path(_suffix_from_uri(uri_s)).suffix.lower() in (".xlsx", ".xlsm"):
      cfg = source_dataset.ingestion_config or {}

      sheet_name = cfg.get("sheet_name")
      sheet_index = cfg.get("sheet_index")

      # Guard: do not allow both sheet_name and sheet_index
      if sheet_name is not None and sheet_index is not None:
        raise ValueError("Specify either 'sheet_name' or 'sheet_index', not both.")

      # Safe casts
      header_row = int(cfg.get("header_row", 1))
      max_rows = cfg.get("max_rows")
      if max_rows is not None:
        max_rows = int(max_rows)
      if sheet_index is not None:
        sheet_index = int(sheet_index)

      records = _load_excel_records_from_bytes(
        _read_bytes(uri_s),
        sheet_name=sheet_name,
        sheet_index=sheet_index,
        header_row=header_row,
        max_rows=max_rows,
      )
    else:
      cfg = source_dataset.ingestion_config or {}
      delimiter = cfg.get("delimiter")
      quotechar = cfg.get("quotechar")
      encoding = cfg.get("encoding")
      records = _load_file_records(
        uri_s,
        file_type=file_type,
        delimiter=delimiter,
        quotechar=quotechar,
        encoding=encoding,
      )

    if not records:
      return {"rows_extracted": 0, "landing": None}
    rows_extracted = len(records)
    landing = land_raw_json_records(
      target_engine=target_engine,
      target_dialect=dialect,
      td=td,
      records=records,
      batch_run_id=batch_run_id,
      load_run_id=load_run_id,
      target_system=target_system,
      profile=profile,
      meta_schema=meta_schema,
      source_system_short_name=str(source_dataset.source_system.short_name),
      source_dataset_name=str(source_dataset.source_dataset_name),
      source_object=uri_s,
      ingest_mode=(file_type or "file"),
      chunk_size=chunk_size,
      source_dataset=source_dataset,
      strict=False,
    )

  return {
    "rows_extracted": rows_extracted,
    # If chunked ingestion accumulated totals, prefer the total in the returned landing dict.
    "landing": ({**(landing or {}), "rows_inserted": rows_inserted_total} if rows_inserted_total > 0 else landing),
  }

"""
elevata - Metadata-driven Data Platform Framework
Copyright © 2025-2026 Ilona Tag

This file is part of elevata.

elevata is free software: you can redistribute it and/or modify
it under the terms of the GNU Affero General Public License as
published by the Free Software Foundation, either version 3 of
the License, or (at your option) any later version.

elevata is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
GNU Affero General Public License for more details.

You should have received a copy of the GNU Affero General Public License
along with elevata. If not, see <https://www.gnu.org/licenses/>.

Contact: <https://github.com/elevata-labs/elevata>.
"""

from __future__ import annotations

import csv
import io
import json
import re
import logging
import os
import urllib.parse
import urllib.request
from typing import Any, Dict, List

from django.db import transaction

from metadata.ingestion.infer import infer_column_profile, infer_pk_columns
from metadata.models import SourceColumn
from metadata.ingestion.normalization import normalize_column_name

log = logging.getLogger(__name__)


def _normalize_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
  """
  Normalize record keys so sampling/inference matches SourceColumn.source_column_name.
  """
  out: List[Dict[str, Any]] = []
  for r in rows:
    rr: Dict[str, Any] = {}
    for k, v in (r or {}).items():
      nk = normalize_column_name(k)
      rr[nk] = v
    out.append(rr)
  return out


def _resolve_uri(ds) -> str:
  cfg = getattr(ds, "ingestion_config", None) or {}
  if not isinstance(cfg, dict):
    cfg = {}
  uri = (cfg.get("uri") or cfg.get("url") or cfg.get("path") or cfg.get("file_path") or "").strip()
  if uri:
    return os.path.expandvars(uri)
  raise ValueError(
    f"File SourceDataset '{ds.source_system.short_name}:{ds.source_dataset_name}' "
    "is missing ingestion_config.uri"
  )


def _read_bytes(uri: str, *, max_bytes: int = 2_000_000) -> bytes:
  uri = os.path.expandvars(uri or "")
  p = urllib.parse.urlparse(uri)
  if p.scheme in ("http", "https"):
    with urllib.request.urlopen(uri, timeout=60) as resp:
      data = resp.read(max_bytes + 1)
    return data[:max_bytes]
  if uri.startswith("file://"):
    uri = uri[len("file://"):]
    # Windows file URI fix: file:///C:/... becomes /C:/... after stripping.
    # Convert /C:/... -> C:/... so os.path.exists works on Windows.
    if re.match(r"^/[A-Za-z]:/", uri):
      uri = uri[1:]
  if not os.path.exists(uri):
    raise FileNotFoundError(uri)
  with open(uri, "rb") as f:
    return f.read(max_bytes)


def _local_path_from_uri(uri: str) -> str:
  """
  Convert file:// URIs to local paths. Keeps plain local paths unchanged.
  Supports both:
    - file:///C:/path/file.xlsx
    - file://C:/path/file.xlsx
  """
  u = (uri or "").strip()
  if u.startswith("file://"):
    u = u[len("file://"):]
    # Windows file URI fix: file:///C:/... becomes /C:/... after stripping.
    # Convert /C:/... -> C:/... so os.path.exists works on Windows.
    if re.match(r"^/[A-Za-z]:/", u):
      u = u[1:]

  return u


def _sample_csv(
  uri: str,
  *,
  max_rows: int = 200,
  delimiter: str | None = None,
  quotechar: str | None = None,
  encoding: str | None = None,
) -> List[Dict[str, Any]]:
  raw = _read_bytes(uri)
  enc = (encoding or "utf-8").strip() or "utf-8"
  txt = raw.decode(enc, errors="replace")

  fh = io.StringIO(txt)
  d = (delimiter or ",")
  qc = (quotechar or '"')
  reader = csv.DictReader(fh, delimiter=d, quotechar=qc)

  rows = []
  for i, r in enumerate(reader):
    if i >= max_rows:
      break
    rows.append({(k or ""): (v if v != "" else None) for k, v in (r or {}).items()})
  return rows


def _sample_json(uri: str, *, max_rows: int = 200) -> List[Dict[str, Any]]:
  raw = _read_bytes(uri)
  payload = json.loads(raw.decode("utf-8", errors="replace"))
  if isinstance(payload, list):
    out = [r for r in payload if isinstance(r, dict)]
    return out[:max_rows]
  if isinstance(payload, dict):
    for k in ("items", "data", "results"):
      v = payload.get(k)
      if isinstance(v, list):
        out = [r for r in v if isinstance(r, dict)]
        return out[:max_rows]
  raise ValueError("Unsupported JSON shape (expected array of objects or dict with items/data/results).")


def _sample_jsonl(uri: str, *, max_rows: int = 200) -> List[Dict[str, Any]]:
  raw = _read_bytes(uri)
  txt = raw.decode("utf-8", errors="replace")
  rows = []
  for line in txt.splitlines():
    line = line.strip()
    if not line:
      continue
    obj = json.loads(line)
    if isinstance(obj, dict):
      rows.append(obj)
    if len(rows) >= max_rows:
      break
  return rows


def _sample_excel(
  uri: str,
  *,
  max_rows: int = 200,
  sheet_name: str | None = None,
  sheet_index: int | None = None,
  header_row: int | None = None,
) -> List[Dict[str, Any]]:
  """
  Sample rows from an .xlsx file using openpyxl.
  - sheet_name: preferred selector (if provided and exists)
  - sheet_index: 0-based index fallback (default 0)
  - header_row: 1-based row index for headers (default 1)
  """
  # Local-only for now (consistent with how file ingestion is typically configured).
  # If you need http(s) Excel sampling later, we can add a temp download path.
  from openpyxl import load_workbook

  path = _local_path_from_uri(uri)
  if not os.path.exists(path):
    raise FileNotFoundError(path)

  wb = load_workbook(filename=path, read_only=True, data_only=True)
  try:
    ws = None
    if sheet_name:
      sn = sheet_name.strip()
      if sn in wb.sheetnames:
        ws = wb[sn]
    if ws is None:
      idx = 0 if sheet_index is None else int(sheet_index)
      idx = max(0, min(idx, len(wb.sheetnames) - 1))
      ws = wb[wb.sheetnames[idx]]

    hr = 1 if header_row is None else int(header_row)
    hr = max(1, hr)

    # Read header row (1-based)
    header_cells = list(ws.iter_rows(min_row=hr, max_row=hr, values_only=True))[0]
    headers = []
    for c in header_cells:
      name = (str(c).strip() if c is not None else "")
      headers.append(name)

    rows: List[Dict[str, Any]] = []
    # Data starts at hr+1
    for i, values in enumerate(ws.iter_rows(min_row=hr + 1, values_only=True)):
      if i >= max_rows:
        break
      if values is None:
        continue
      rec: Dict[str, Any] = {}
      for j, h in enumerate(headers):
        if not h:
          continue
        v = values[j] if j < len(values) else None
        rec[h] = v
      # Skip completely empty records
      if any(v is not None and v != "" for v in rec.values()):
        rows.append(rec)
    return rows
  finally:
    try:
      wb.close()
    except Exception:
      pass


def _sample_parquet(
  uri: str,
  *,
  max_rows: int = 200,
) -> list[dict]:
  """
  Sample rows from a parquet file using pyarrow.
  Returns list of dict records.
  """
  path = _local_path_from_uri(uri)
  if not os.path.exists(path):
    raise FileNotFoundError(path)

  import pyarrow.parquet as pq

  pf = pq.ParquetFile(path)
  # Read only a small number of rows
  table = pf.read_row_groups([0]) if pf.num_row_groups > 0 else pf.read()
  if table.num_rows > max_rows:
    table = table.slice(0, max_rows)

  # Convert to python dict-of-lists, then to list-of-dicts
  cols = table.column_names
  data = table.to_pydict()
  rows: list[dict] = []
  for i in range(table.num_rows):
    r = {}
    for c in cols:
      r[c] = data[c][i]
    rows.append(r)
  return rows


def import_file_metadata_for_dataset(
  ds,
  *,
  file_type: str,
  autointegrate_pk: bool = True,
  reset_flags: bool = False,
) -> Dict[str, Any]:
  uri = _resolve_uri(ds)
  ft = (file_type or "").lower()
  uri_l = (uri or "").lower()
  if ft == "json" and (uri_l.endswith(".jsonl") or uri_l.endswith(".ndjson")):
    ft = "jsonl"

  cfg = getattr(ds, "ingestion_config", None) or {}
  if not isinstance(cfg, dict):
    cfg = {}

  if ft == "csv":
    rows = _sample_csv(
      uri,
      delimiter=(cfg.get("delimiter") or None),
      quotechar=(cfg.get("quotechar") or None),
      encoding=(cfg.get("encoding") or None),
    )

  elif ft == "json":
    rows = _sample_json(uri)
  elif ft == "jsonl":
    rows = _sample_jsonl(uri)
  elif ft == "excel":
    rows = _sample_excel(
      uri,
      sheet_name=(cfg.get("sheet_name") or None),
      sheet_index=(cfg.get("sheet_index") if "sheet_index" in cfg else None),
      header_row=(cfg.get("header_row") if "header_row" in cfg else None),
    )
  elif ft == "parquet":
    rows = _sample_parquet(uri)

  else:
    raise NotImplementedError(f"Unsupported file type for auto import: '{file_type}'")

  if not rows:
    raise ValueError(
      f"No rows found for file dataset '{ds.source_system.short_name}:{ds.source_dataset_name}' (uri={uri!r})."
    )

  rows = _normalize_rows(rows)

  keys = set()
  for r in rows:
    keys.update(r.keys())
  col_names = [normalize_column_name(k) for k in sorted(keys) if k]

  pk_final = set(infer_pk_columns(rows, col_names))

  existing: Dict[str, SourceColumn] = {c.source_column_name: c for c in ds.source_columns.all()}
  seen = set()

  created = 0
  updated = 0
  removed = 0

  with transaction.atomic():
    if reset_flags:
      ds.source_columns.update(
        integrate=False,
        pii_level="none",
        description="",
        primary_key_column=False,
      )

    if existing:
      base = 10000
      n = 0
      for sc0 in existing.values():
        n += 1
        sc0.ordinal_position = base + n
        sc0.save(update_fields=["ordinal_position"])

    for i, col in enumerate(col_names, start=1):
      sc = existing.get(col)
      is_new = sc is None
      if sc is None:
        sc = SourceColumn(
          source_dataset=ds,
          source_column_name=col,
          integrate=True,
          pii_level="none",
        )
        created += 1

      sc.ordinal_position = i
      sc.source_datatype_raw = None

      values = [r.get(col) for r in rows if col in r]
      dtype, max_len, dec_prec, dec_scale = infer_column_profile(values)
      sc.datatype = dtype
      sc.max_length = max_len
      sc.decimal_precision = dec_prec
      sc.decimal_scale = dec_scale
      sc.nullable = True
      sc.primary_key_column = col in pk_final
      sc.referenced_source_dataset_name = None
      sc.json_path = f"$.{col}"

      if reset_flags and is_new is False:
        sc.integrate = False
      if autointegrate_pk and sc.primary_key_column:
        sc.integrate = True

      sc.save()
      if not is_new:
        updated += 1
      seen.add(col)

    to_remove = [c for name, c in existing.items() if name not in seen]
    if to_remove:
      removed = len(to_remove)
      SourceColumn.objects.filter(pk__in=[c.pk for c in to_remove]).delete()

  return {
    "columns_imported": len(seen),
    "created": created,
    "updated": updated,
    "removed": removed,
    "pk_detected": sorted(pk_final),
  }
